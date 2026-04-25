"""Post-simulation artifact: render the residuals CSV into a styled XLSX with a log-scale chart."""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


FONT = "Arial"

_HEADER_FILL = PatternFill("solid", start_color="FF1F3A5F")
_HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFFFF", size=11)
_TITLE_FONT = Font(name=FONT, bold=True, size=14, color="FF1F3A5F")
_LABEL_FONT = Font(name=FONT, bold=True, size=11, color="FF334155")
_VALUE_FONT = Font(name=FONT, size=11)
_THIN = Side(style="thin", color="FFCBD5E1")
_CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BAND_FILL = PatternFill("solid", start_color="FFF1F5F9")


def write_residuals_xlsx(csv_path: Path, xlsx_path: Path, meta: dict) -> None:
    """Build a workbook with a Run Info sheet and a Residuals sheet + log-scale line chart."""
    csv_path = Path(csv_path)
    xlsx_path = Path(xlsx_path)

    rows = []
    with csv_path.open("r", encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        for r in reader:
            if len(r) < 3:
                continue
            try:
                rows.append((int(r[0]), float(r[1]), float(r[2])))
            except ValueError:
                continue
    if not rows:
        return

    wb = Workbook()

    info = wb.active
    info.title = "Run Info"
    info["A1"] = "CFD simulation — run info"
    info["A1"].font = _TITLE_FONT
    info.merge_cells("A1:B1")

    items = [
        ("Simulation ID",   meta.get("slug", "")),
        ("Started",         meta.get("started", "")),
        ("Mode",            meta.get("mode", "")),
        ("Reynolds number", meta.get("Re", "")),
        ("Grid (nx × ny)",  f'{meta.get("nx", "")} × {meta.get("ny", "")}'),
        ("Domain (W × H)",  f'{meta.get("width", "")} × {meta.get("height", "")}'),
        ("Iterations",      meta.get("iters", "")),
        ("Converged",       meta.get("converged", "")),
        ("Elapsed (s)",     meta.get("elapsed_s", "")),
        ("Residual target", meta.get("res_drop", "")),
        ("Log file",        meta.get("log_file", "")),
        ("Residuals CSV",   csv_path.name),
    ]
    for i, (k, v) in enumerate(items, start=3):
        info.cell(row=i, column=1, value=k).font = _LABEL_FONT
        c = info.cell(row=i, column=2, value=v)
        c.font = _VALUE_FONT
        c.alignment = Alignment(horizontal="left")
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 42

    ws = wb.create_sheet("Residuals")
    headers = ["Iteration", "Momentum residual (R_mom)", "Continuity residual (R_div)"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _CELL_BORDER

    for i, (it, rm, rd) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=it).font = _VALUE_FONT
        ws.cell(row=i, column=2, value=rm).font = _VALUE_FONT
        ws.cell(row=i, column=3, value=rd).font = _VALUE_FONT
        ws.cell(row=i, column=2).number_format = "0.00E+00"
        ws.cell(row=i, column=3).number_format = "0.00E+00"
        if i % 2 == 0:
            for j in range(1, 4):
                ws.cell(row=i, column=j).fill = _BAND_FILL

    last = len(rows) + 1
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 28
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    chart = _build_residuals_chart(ws, rows_last=last, n_rows=len(rows))
    ws.add_chart(chart, "E2")

    wb.save(xlsx_path)


def _solid_line(hex_rgb: str, width_emu: int = 22000) -> GraphicalProperties:
    gp = GraphicalProperties()
    gp.line = LineProperties(solidFill=hex_rgb, w=width_emu)
    return gp


def _build_residuals_chart(ws, *, rows_last: int, n_rows: int) -> LineChart:
    chart = LineChart()
    chart.title = "Residual convergence history"
    chart.y_axis.title = "Residual (log scale)"
    chart.x_axis.title = "Iteration"
    chart.legend.position = "b"
    chart.height = 14
    chart.width = 26

    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=rows_last)
    cats = Reference(ws, min_col=1, min_row=2, max_row=rows_last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    colors = ("FF1F3A5F", "FFB91C1C")
    for series, color in zip(chart.series, colors):
        series.smooth = False
        series.graphicalProperties = _solid_line(color, width_emu=22000)

    chart.y_axis.axPos = "l"
    chart.y_axis.scaling.logBase = 10
    chart.y_axis.crosses = "autoZero"
    chart.x_axis.axPos = "b"

    return chart
